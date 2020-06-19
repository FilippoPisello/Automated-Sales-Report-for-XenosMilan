#!/usr/bin/env python
#Created by Filippo Pisello

def formatting(file_name, table_index, table_columns, index=False,
               format_header=True, format_index_main=False,
               format_index_light=True,
               color="0066cc", color_col="0066cc", color_light="b2beb5",
               custom_width=20):
    """
    Take a raw excel created with pandas and apply formatting

    ---------------
    Variables description:
    file_name: name of the original file, extension included (ex: "house.xlsx")
    table_index: name of the index from pandas (ex: df.index)
    table_columns: same with columns
    index: tells if the table has/was exported with an index, it affects what is
    formatted how
    format_header: tells if the header should be formatted
    format_index_main: tells if the index/first column should be formatted as
    heavy as the header
    format_index_light: tells if the index/first column should be formatted with
    a lighter formatting
    """

    from openpyxl import load_workbook
    from openpyxl.styles import Font, Color, Alignment, Border, Side, colors, PatternFill
    import string

    def find_header_and_index(table_index, table_columns, index=False):
        dim_columns = len(table_columns)
        dim_index = len(table_index)

        #Header
        header = []    # A list containing pairs in style "A1, B2..."
        if index:
            for number in range(0, dim_columns):
                header.append(str(string.ascii_uppercase[number + 1] + "1"))
        else:
            for number in range(0, dim_columns):
                header.append(str(string.ascii_uppercase[number] + "1"))

        #Index
        index = []
        for number in range(0, dim_index):
            index.append("A" + str(number + 2))

        #Body
        body = []
        if index:
            start = 1
        else:
            start = 0
        for value in range(0, dim_columns):
            letter = string.ascii_uppercase[value]
            for number in range(0, dim_index):
                body.append(str(letter + str(number + 2)))

        return header, index, body

    header, index, body = find_header_and_index(table_index = table_index,
                                                table_columns = table_columns,
                                                index=index)
    file_name = str(file_name)
    workbook = load_workbook(filename = file_name)
    sheet = workbook.active
    header_font = Font( bold=True, color="ffffff", size=12)
    first_column_font = Font(bold=True)
    center_aligned_text = Alignment(horizontal="center")
    left_aligned_text = Alignment(horizontal="left")
    right_aligned_text = Alignment(horizontal="right")
    header_fill = PatternFill(start_color=color, end_color=color,
                              fill_type="solid")
    index_fill = PatternFill(start_color=color_col, end_color=color_col,
                             fill_type="solid")
    light_fill = PatternFill(start_color=color_light, end_color=color_light,
                             fill_type="solid")

    if format_header:
        for cell in header:
            sheet[cell].font = header_font
            sheet[cell].fill = header_fill
            sheet[cell].alignment = center_aligned_text

    if format_index_main:
        for cell in index:
            sheet[cell].font = header_font
            sheet[cell].fill = index_fill
            sheet[cell].alignment = center_aligned_text

    if format_index_light:
        for cell in index:
            sheet[cell].font = first_column_font
            sheet[cell].fill = light_fill
            sheet[cell].alignment = left_aligned_text

    for cell in body:
        sheet[cell].alignment = right_aligned_text

    for value in range(0, len(table_columns)+1):
        sheet.column_dimensions[string.ascii_uppercase[value]].width = custom_width

    workbook.save(filename="formatted_{}".format(file_name))


#TO TEST (requires a prova.xlsx file)
#col = ["A", "B", "A", "B", "A", "B"]
#ind = [1,2,3,4,5,6,7,8,9,10]
#formatting(prova.xlsx, ind, col, index=True)
